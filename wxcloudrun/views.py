import json
import logging
import time
from django.http import JsonResponse
from django.shortcuts import render
import requests
import os
import openpyxl

logger = logging.getLogger('log')
# Load the Excel table once and keep it in memory (optional optimization)


def download_excel_file(temp_url, save_path="/tmp/rules.xlsx"):
    response = requests.get(temp_url)
    if response.status_code == 200:
        with open(save_path, "wb") as f:
            f.write(response.content)
        return save_path
    else:
        raise Exception(f"Failed to download file: {response.status_code}")


def load_table(path):
    if not os.path.exists(path):
        return {}, [], [], "文件不存在", "文件不存在"
    wb = openpyxl.load_workbook(path)
    sheet1 = wb.worksheets[0]
    sheet2 = wb.worksheets[1]

    # Load default fallback message from Sheet 2, cell A1
    default_message = sheet2.cell(row=1, column=1).value or "No default message set."
    warning_message = sheet2.cell(row=2, column=1).value or "No warning message set."

    data = {}
    headers = [cell.value for cell in sheet1[2]][1:]  # First row, skip first cell
    count = len(headers)  # count of info-types, including 3 jurisdiction columns

    # for row in sheet1.iter_rows(min_row=2):
    #    jurisdiction = row[0].value
    #    if not jurisdiction:
    #        continue
    #    data[jurisdiction] = {}
    #    for i, cell in enumerate(row[1:]):
    #        data[jurisdiction][headers[i]] = cell.value if cell.value else 'No data available.'
    for row in sheet1.iter_rows(min_row=2):
        jurisdiction = row[1].value  # Column B
        if not jurisdiction:
            continue
        data[jurisdiction] = {}
        for i in range(4, count):  # Columns E (4) to end
            header = headers[i - 4]  # headers aligned with E to end
            cell = row[i]
            data[jurisdiction][header] = cell.value if cell.value else 'No data available.'
        jurisdiction = row[2].value  # Column C
        if not jurisdiction:
            continue
        data[jurisdiction] = {}
        for i in range(4, count):  # Columns E (4) to end
            header = headers[i - 4]  # headers aligned with E to end
            cell = row[i]
            data[jurisdiction][header] = cell.value if cell.value else 'No data available.'
        jurisdiction = row[3].value  # Column D
        if not jurisdiction:
            continue
        data[jurisdiction] = {}
        for i in range(4, count):  # Columns E (4) to end
            header = headers[i - 4]  # headers aligned with E to end
            cell = row[i]
            data[jurisdiction][header] = cell.value if cell.value else 'No data available.'
    return data, list(data.keys()), headers, default_message, warning_message

def fuzzy_match(text, options):
    text = str(text).lower()
    for option in options:
        if option and str(option).lower() in text.lower():
            return option
    return None


def test(request):
    # load the file to local path
    file_url = (
        "https://7072-prod-1g3d62ey10e2634f-1353111496.tcb.qcloud.la/rules.xlsx"
        "?sign=4fabc22871e948dbff37604a24e2f2f7&t=1744858377"
    )
    local_path = download_excel_file(file_url)

    # 1. load the table to memory
    table, jurisdictions, info_types, default_message, warning_message = load_table(local_path)

    #  用于微信系统健康检查，在系统向服务发送GET时返回ok
    if request.method == 'GET' or request.method == 'get':
        return JsonResponse({'code': 0, 'msg': 'ok'}, json_dumps_params={'ensure_ascii': False})
    try:
        # 2. 获取请求内容
        body_unicode = request.body.decode('utf-8')
        body = json.loads(body_unicode)  # 转换为JSON格式
        logger.info('解析后的消息内容：{}'.format(body))
    except Exception as e:
        logger.error('解析请求失败: {}'.format(str(e)))
        return JsonResponse({'code': -1, 'errorMsg': 'JSON解析失败'}, json_dumps_params={'ensure_ascii': False})

    # 3. 从请求中获取消息内容
    user_msg = body.get('Content', '')  # 用户输入的消息
    from_user = body.get('FromUserName', '')  # 用户的公众号ID
    to_user = body.get('ToUserName', '')  # 公众号ID

    if not user_msg:
        reply = default_message
    elif "留言" in user_msg:
        reply = "我们已收到您的留言，我们会尽快回复~"
    else:
        matched_jurisdiction = fuzzy_match(user_msg, jurisdictions)
        if not matched_jurisdiction:
            reply = default_message
        else:
            matched_info = fuzzy_match(user_msg, info_types)
            if not matched_info:
                reply = default_message
            else:
                result = table[matched_jurisdiction][matched_info]
                reply = warning_message + "\n" + matched_jurisdiction + "-" + matched_info + ":\n" + result

    # 5. 构建回复格式
    rsp = JsonResponse({
            "ToUserName": from_user,
            "FromUserName": to_user,
            "CreateTime": int(time.time()),
            "MsgType": "text",
            "Content": reply,
    }, json_dumps_params={'ensure_ascii': False})

    logger.info('response result: {}'.format(rsp.content.decode('utf-8')))
    return rsp
